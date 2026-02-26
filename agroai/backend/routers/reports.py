"""AgroAI — Reports Router (PDF / Excel)"""
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
import io
from datetime import date

from models.database import get_db, Livestock, AIDetection, Farm
from routers.auth import get_current_user, User

router = APIRouter()


@router.get("/pdf")
async def generate_pdf_report(
    farm_id: int = 1,
    period: str = "monthly",
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """PDF есеп жасау"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # Header
        elements.append(Paragraph(f"AgroAI — Мал шаруашылығы есебі", styles["Title"]))
        elements.append(Paragraph(f"Жасалған: {date.today()}", styles["Normal"]))
        elements.append(Spacer(1, 12))

        # Stats
        total_res = await db.execute(select(func.count(Livestock.id)))
        total = total_res.scalar() or 0

        data = [
            ["Көрсеткіш", "Мән"],
            ["Жалпы мал саны", str(total)],
            ["Ферма ID", str(farm_id)],
            ["Есеп кезеңі", period],
            ["Жасалған күні", str(date.today())],
        ]
        table = Table(data)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2e7d32")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 12),
            ("GRID", (0, 0), (-1, -1), 1, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f8e9")]),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=agroai_report_{date.today()}.pdf"},
        )
    except ImportError:
        return {"error": "reportlab not installed", "message": "pip install reportlab"}


@router.get("/excel")
async def generate_excel_report(
    farm_id: int = 1,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Excel есеп жасау"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Мал тізімі"

        # Headers
        headers = ["ID", "Тег №", "Тұқымы", "Жынысы", "Салмақ (кг)", "Денсаулық", "Тіркелген"]
        header_fill = PatternFill(start_color="2e7d32", end_color="2e7d32", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data
        result = await db.execute(select(Livestock).limit(1000))
        livestock_list = result.scalars().all()

        for row, animal in enumerate(livestock_list, 2):
            ws.cell(row=row, column=1, value=animal.id)
            ws.cell(row=row, column=2, value=animal.tag_number or "—")
            ws.cell(row=row, column=3, value=animal.breed or "—")
            ws.cell(row=row, column=4, value=animal.sex or "—")
            ws.cell(row=row, column=5, value=animal.weight_kg)
            ws.cell(row=row, column=6, value=str(animal.health_status))
            ws.cell(row=row, column=7, value=str(animal.created_at)[:10] if animal.created_at else "—")

        # Auto column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=agroai_livestock_{date.today()}.xlsx"},
        )
    except ImportError:
        return {"error": "openpyxl not installed", "message": "pip install openpyxl"}
